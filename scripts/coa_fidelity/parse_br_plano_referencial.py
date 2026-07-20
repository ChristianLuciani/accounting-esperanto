#!/usr/bin/env python3
"""
Deterministic parser for Brazil's "Plano de Contas Referencial" (Receita
Federal do Brasil / RFB, published via the SPED ECF Leiaute 12 dynamic-tables
workbook -- registers L100A/L300A cover the "PJ do Lucro Real - PJ em Geral"
profile referenced by ECF register J051). Structurally different from both
Ecuador's "CODIGO / NOMBRE / SIGNO" text-table layout and Mexico's
"Nivel / Código agrupador / Nombre" text-table layout (both extracted from
pdftotext output) -- Brazil's source is a well-structured .xlsx workbook with
explicit columns, so it gets its own parser reading directly via openpyxl
rather than being forced through a text-table regex.

Unlike EC/MX, Brazil's own data already distinguishes header/aggregate rows
from postable leaves via an authoritative TIPO column ("S" = Sintética,
header/subtotal; "A" = Analítica, postable leaf) and an explicit CONTA
SUPERIOR (parent code) column -- no prefix-heuristic subtotal-guessing is
needed the way EC/MX's SUBTOTAL_PREFIXES tables were needed. This parser
carries TIPO and the parent code through verbatim so map_official_chart.py
can use them directly instead of re-deriving hierarchy from string prefixes.

This exists so that transcribing the official chart into Kontablo never goes
through hand-typing -- the same failure mode that produced Ecuador's original
13-account draft and Mexico's original 14-account draft against the real
charts of 721 and 1,079 codes respectively. The input is the RFB-published
"Tabelas_Dinamicas_ECF_Leiaute_12*.xlsx" workbook (downloaded from
http://sped.rfb.gov.br/arquivo/download/8002 -- note /download/, not
/show/, which is only an HTML landing page); the output is a raw, verbatim,
source-cited YAML -- no Kontablo UUID classification happens here, that is a
separate, reviewable step (see map_official_chart.py).

Usage:
    python3 scripts/coa_fidelity/parse_br_plano_referencial.py \
        --input Tabelas_ECF_Leiaute12.xlsx \
        --source-url "http://sped.rfb.gov.br/arquivo/download/8002" \
        --out localizations/br/plano_referencial_official_chart.yaml
"""
import argparse
import re
import sys

import openpyxl
import yaml

# The two sheets covering the general-commercial-entity profile ("PJ do
# Lucro Real - PJ em Geral") -- the standard profile analogous to what EC's
# (Supercias NIIF) and MX's (SAT Código Agrupador general regime) rounds
# cover. L100B/L100C and L300B/L300C (financial institutions / insurers) and
# the P100A/P300 set (Lucro Presumido regime) are deliberately out of scope
# for this round -- documented in localizations/br/README.md, not silently
# dropped.
SHEETS = {
    "L100A": "Balanço Patrimonial (Contas Patrimoniais) - PJ do Lucro Real - PJ em Geral",
    "L300A": "Demonstração do Resultado (Contas de Resultado) - PJ do Lucro Real - PJ em Geral",
}

NATUREZA_LABEL = {
    1: "Ativo",
    2: "Passivo",
    3: "Patrimônio Líquido",
    4: "Resultado",
}

CONTRA_PREFIX_RE = re.compile(r"^\(\s*-\s*\)\s*")


def parse_sheet(ws, sheet_name, statement_label):
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    expected = ["CÓDIGO", "DESCRIÇÃO", "DT_INI", "DT_FIM", "TIPO",
                "CONTA SUPERIOR", "NÍVEL", "NATUREZA", "ORIENTAÇÕES"]
    for col in expected:
        if col not in header:
            print(f"ERROR: sheet {sheet_name} missing expected column {col!r} "
                  f"(found {header})", file=sys.stderr)
            sys.exit(1)
    idx = {col: header.index(col) for col in expected}

    accounts = []
    for r in rows[1:]:
        code_raw = r[idx["CÓDIGO"]]
        if code_raw is None:
            continue
        code = str(code_raw).strip()
        name_raw = (r[idx["DESCRIÇÃO"]] or "").strip()
        tipo = (r[idx["TIPO"]] or "").strip()
        superior_raw = r[idx["CONTA SUPERIOR"]]
        superior = str(superior_raw).strip() if superior_raw is not None else None
        nivel = r[idx["NÍVEL"]]
        natureza = r[idx["NATUREZA"]]
        orientacoes = r[idx["ORIENTAÇÕES"]]
        orientacoes = orientacoes.strip() if isinstance(orientacoes, str) else None

        is_contra = bool(CONTRA_PREFIX_RE.match(name_raw))
        name_clean = CONTRA_PREFIX_RE.sub("", name_raw).strip()

        accounts.append({
            "code": code,
            "name": name_clean,
            "name_raw": name_raw,
            "tipo": tipo,               # "S" (Sintética/header) or "A" (Analítica/leaf) -- authoritative, from source
            "conta_superior": superior,  # explicit parent code from source, or null for a root
            "nivel": int(nivel) if nivel is not None else None,
            "natureza": int(natureza) if natureza is not None else None,
            "is_contra": is_contra,      # source marks contra-accounts with a literal "(-)" name prefix
            "orientacoes": orientacoes,
            "statement": statement_label,
            "source_sheet": sheet_name,
        })
    return accounts


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="RFB Tabelas_Dinamicas_ECF xlsx workbook")
    ap.add_argument("--source-url", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=True, read_only=True)

    all_accounts = []
    for sheet_name, statement_label in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"ERROR: expected sheet {sheet_name!r} not found in workbook "
                  f"(available: {wb.sheetnames})", file=sys.stderr)
            sys.exit(1)
        accounts = parse_sheet(wb[sheet_name], sheet_name, statement_label)
        if not accounts:
            print(f"ERROR: no account rows parsed from sheet {sheet_name}", file=sys.stderr)
            sys.exit(1)
        all_accounts.append((sheet_name, accounts))

    codes_seen = set()
    duplicates = []
    flat = []
    for _, accounts in all_accounts:
        for a in accounts:
            if a["code"] in codes_seen:
                duplicates.append(a["code"])
            codes_seen.add(a["code"])
            flat.append(a)

    doc = {
        "metadata": {
            "jurisdiction": "br",
            "authority": "Receita Federal do Brasil (RFB) / SERPRO -- SPED (Sistema Público de "
                         "Escrituração Digital)",
            "source_url": args.source_url,
            "source_document": "Tabelas Dinâmicas e Planos de Contas Referenciais - Leiaute 12 da "
                                "ECF (Escrituração Contábil Fiscal), Anexo ao Ato Declaratório "
                                "Executivo Cofis nº 02/2026, atualização de 28/05/2026",
            "profile": "PJ do Lucro Real - PJ em Geral (general commercial/industrial entity under "
                       "the Lucro Real tax regime) -- registers L100A (Contas Patrimoniais / balance "
                       "sheet) and L300A (Contas de Resultado / P&L). L100B/L100C and L300B/L300C "
                       "(financial institutions, insurers) and the P100A/P300 Lucro Presumido set "
                       "are out of scope for this round.",
            "extraction_command": "python3 scripts/coa_fidelity/parse_br_plano_referencial.py "
                                   "--input Tabelas_ECF_Leiaute12.xlsx --source-url <URL> --out <OUT>",
            "parser": "scripts/coa_fidelity/parse_br_plano_referencial.py",
            "total_accounts": len(flat),
            "sheets": {sn: len(accounts) for sn, accounts in all_accounts},
            "duplicate_codes": sorted(set(duplicates)) or None,
            "natureza_legend": NATUREZA_LABEL,
            "note": (
                "Verbatim transcription of the primary-source official Plano de Contas "
                "Referencial (RFB SPED ECF Leiaute 12, L100A + L300A sheets). No Kontablo "
                "UUID classification here by design -- that mapping is a separate, "
                "human-reviewable step (map_official_chart.py) so a parsing bug can never "
                "silently corrupt the semantic mapping. TIPO ('S'=Sintética/header, "
                "'A'=Analítica/leaf) and CONTA SUPERIOR (explicit parent code) are carried "
                "through verbatim from the source -- Brazil's own data already distinguishes "
                "header/aggregate rows from postable leaves authoritatively, unlike EC/MX "
                "where this had to be inferred from a prefix heuristic. Account names "
                "beginning with a literal '(-)' marker in the source are contra-accounts "
                "(allowances, accumulated depreciation/amortization/exhaustion, sales/"
                "purchase deductions, impairment losses); is_contra=true carries this "
                "forward and name has the '(-)' prefix stripped (name_raw keeps the original)."
            ),
        },
        "accounts": flat,
    }

    with open(args.out, "w", encoding="utf-8") as f:
        yaml.safe_dump(doc, f, allow_unicode=True, sort_keys=False, width=100)

    print(f"Parsed {len(flat)} accounts ({len(codes_seen)} unique codes)")
    for sn, accounts in all_accounts:
        print(f"  {sn}: {len(accounts)}")
    if duplicates:
        print(f"WARNING: {len(set(duplicates))} duplicate codes: {sorted(set(duplicates))[:10]}...")
    print(f"Written: {args.out}")


if __name__ == "__main__":
    main()
